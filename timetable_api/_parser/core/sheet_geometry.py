from __future__ import annotations

from datetime import datetime, timedelta

from openpyxl.utils import get_column_letter
from openpyxl.worksheet.worksheet import Worksheet

from timetable_api._parser.core.models import CellBounds, RawCell
from timetable_api._parser.extractors.day_slots import Slot


def build_merge_bounds(sheet: Worksheet) -> dict[tuple[int, int], CellBounds]:
    merge_bounds: dict[tuple[int, int], CellBounds] = {}
    for merged_range in sheet.merged_cells.ranges:
        bounds = CellBounds(
            min_row=merged_range.min_row,
            min_col=merged_range.min_col,
            max_row=merged_range.max_row,
            max_col=merged_range.max_col,
        )
        for row in range(merged_range.min_row, merged_range.max_row + 1):
            for col in range(merged_range.min_col, merged_range.max_col + 1):
                merge_bounds[(row, col)] = bounds
    return merge_bounds


def visible_bounds_for_cell(merge_bounds: dict[tuple[int, int], CellBounds], row: int, col: int) -> CellBounds:
    return merge_bounds.get((row, col), CellBounds(min_row=row, min_col=col, max_row=row, max_col=col))


def raw_cells_in_bounds(sheet: Worksheet, min_row: int, max_row: int, bounds: CellBounds) -> list[RawCell]:
    cells: list[RawCell] = []
    seen_coordinates: set[str] = set()

    for row in range(min_row, max_row + 1):
        for col in range(bounds.min_col, bounds.max_col + 1):
            value = clean_text(sheet.cell(row=row, column=col).value)
            if value is None:
                continue

            coordinate = f"{get_column_letter(col)}{row}"
            if coordinate in seen_coordinates:
                continue

            seen_coordinates.add(coordinate)
            cells.append(RawCell(coordinate=coordinate, row=row, column=col, value=value))

    return cells


def rectangle_has_raw(sheet: Worksheet, min_row: int, max_row: int, bounds: CellBounds) -> bool:
    return bool(raw_cells_in_bounds(sheet, min_row, max_row, bounds))


def row_has_top_border(sheet: Worksheet, row: int, bounds: CellBounds) -> bool:
    for col in range(bounds.min_col, bounds.max_col + 1):
        border = sheet.cell(row=row, column=col).border.top
        if border and border.style:
            return True
    return False


def row_has_bottom_border(sheet: Worksheet, row: int, bounds: CellBounds) -> bool:
    for col in range(bounds.min_col, bounds.max_col + 1):
        border = sheet.cell(row=row, column=col).border.bottom
        if border and border.style:
            return True
    return False


def malformed_excel_detail(
    sheet: Worksheet,
    bounds: CellBounds,
    merge_bounds: dict[tuple[int, int], CellBounds],
    periods: int,
) -> str | None:
    issues: list[str] = []
    merged = False
    crosses_merge = False

    for row in range(bounds.min_row, bounds.max_row + 1):
        for col in range(bounds.min_col, bounds.max_col + 1):
            merged_bounds = merge_bounds.get((row, col))
            if merged_bounds is None:
                continue
            merged = True
            if not contains(bounds, merged_bounds):
                crosses_merge = True

    edge_count = sum(
        (
            row_has_top_border(sheet, bounds.min_row, bounds),
            row_has_bottom_border(sheet, bounds.max_row, bounds),
            column_has_left_border(sheet, bounds.min_col, bounds),
            column_has_right_border(sheet, bounds.max_col, bounds),
        )
    )

    if crosses_merge:
        issues.append("the extracted bounds cut through a merged range")
    if not merged and edge_count == 0:
        issues.append("the block has no detectable merged range or outer border")
    if periods > 4 and not row_has_bottom_border(sheet, bounds.max_row, bounds):
        issues.append("the long block has no closing bottom border")

    return "; ".join(issues) or None


def column_has_left_border(sheet: Worksheet, col: int, bounds: CellBounds) -> bool:
    return any(
        sheet.cell(row=row, column=col).border.left.style
        for row in range(bounds.min_row, bounds.max_row + 1)
    )


def column_has_right_border(sheet: Worksheet, col: int, bounds: CellBounds) -> bool:
    return any(
        sheet.cell(row=row, column=col).border.right.style
        for row in range(bounds.min_row, bounds.max_row + 1)
    )


def contains(outer: CellBounds, inner: CellBounds) -> bool:
    return (
        outer.min_row <= inner.min_row
        and outer.min_col <= inner.min_col
        and outer.max_row >= inner.max_row
        and outer.max_col >= inner.max_col
    )


def slot_end_row(slot: Slot, remaining_slots: list[Slot], offset: int) -> int:
    if offset + 1 < len(remaining_slots):
        return remaining_slots[offset + 1].cell.row - 1
    return slot.cell.row + 1


def dedupe_raw_cells(cells: list[RawCell]) -> list[RawCell]:
    seen: set[str] = set()
    unique: list[RawCell] = []
    for cell in sorted(cells, key=lambda item: (item.row, item.column)):
        if cell.coordinate in seen:
            continue
        seen.add(cell.coordinate)
        unique.append(cell)
    return unique


def dedupe_values(values: object) -> list[str]:
    seen: set[str] = set()
    unique: list[str] = []
    for value in values:
        if not isinstance(value, str) or value in seen:
            continue
        seen.add(value)
        unique.append(value)
    return unique


def clean_text(value: object) -> str | None:
    if value is None:
        return None
    text = str(value).strip()
    return text or None


def end_time(start_time: str, periods: int) -> str:
    parsed = datetime.strptime(start_time, "%H:%M")
    return (parsed + timedelta(minutes=50 * periods)).strftime("%H:%M")
