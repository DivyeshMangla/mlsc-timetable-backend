"""Day and slot extraction used only during API startup."""

from __future__ import annotations

from dataclasses import dataclass, field

from openpyxl.cell.cell import Cell
from openpyxl.worksheet.worksheet import Worksheet


SLOT_TIME_BY_SR_NO: dict[int, str] = {
    1: "08:00",
    2: "08:50",
    3: "09:40",
    4: "10:30",
    5: "11:20",
    6: "12:10",
    7: "13:00",
    8: "13:50",
    9: "14:40",
    10: "15:30",
    11: "16:20",
    12: "17:10",
    13: "18:00",
    14: "18:50",
}

DAY_ORDER = ("MONDAY", "TUESDAY", "WEDNESDAY", "THURSDAY", "FRIDAY", "SATURDAY")


@dataclass(frozen=True)
class Slot:
    sr_no: int
    time: str
    cell: Cell


@dataclass
class DaySchedule:
    day: str
    slots: dict[int, Slot] = field(default_factory=dict)


class DaySlotExtractor:
    """Build Day -> Slot -> Cell from the SR NO column."""

    @classmethod
    def extract(cls, sheet: Worksheet) -> dict[str, DaySchedule]:
        sr_no_cell = cls.find_sr_no_cell(sheet)
        if sr_no_cell is None:
            return {}

        schedules: dict[str, DaySchedule] = {}
        day_index = -1
        active_day: DaySchedule | None = None

        for row in range(sr_no_cell.row + 1, sheet.max_row + 1):
            cell = sheet.cell(row=row, column=sr_no_cell.column)
            sr_no = cls.parse_sr_no(cell.value)
            if sr_no is None:
                continue

            if sr_no == 1:
                day_index += 1
                if day_index >= len(DAY_ORDER):
                    break
                day = DAY_ORDER[day_index]
                active_day = schedules.setdefault(day, DaySchedule(day=day))

            if active_day is None or sr_no not in SLOT_TIME_BY_SR_NO:
                continue

            active_day.slots[sr_no] = Slot(
                sr_no=sr_no,
                time=SLOT_TIME_BY_SR_NO[sr_no],
                cell=cell,
            )

        return schedules

    @staticmethod
    def find_sr_no_cell(sheet: Worksheet) -> Cell | None:
        for row in sheet.iter_rows():
            for cell in row:
                if normalize_header(cell.value) in {"SRNO", "SR.NO"}:
                    return cell
                if normalize_header(cell.value) in {"HOUR", "HOURS"}:
                    for offset in range(1, min(cell.column, 4)):
                        slot_header = sheet.cell(row=cell.row, column=cell.column - offset)
                        first_slot = sheet.cell(row=cell.row + 1, column=cell.column - offset)
                        if DaySlotExtractor.parse_sr_no(first_slot.value) == 1:
                            return slot_header
        return None

    @staticmethod
    def parse_sr_no(value: object) -> int | None:
        if isinstance(value, int) and 1 <= value <= 14:
            return value
        if isinstance(value, float) and value.is_integer() and 1 <= value <= 14:
            return int(value)
        if isinstance(value, str):
            text = value.strip()
            if text.isdigit() and 1 <= int(text) <= 14:
                return int(text)
        return None


def normalize_header(value: object) -> str:
    if value is None:
        return ""
    return str(value).strip().upper().replace(" ", "")
