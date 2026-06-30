from __future__ import annotations

from collections.abc import Sequence

from openpyxl.utils import get_column_letter

from timetable_api._parser.core.confidence import ConfidenceReason
from timetable_api._parser.core.models import ClassBlock


def class_blocks_to_jsonable(blocks: dict[str, dict[str, list[ClassBlock]]]) -> dict[str, dict[str, list[dict[str, object]]]]:
    return {
        batch: {
            day: [
                {
                    "start_time": block.start_time,
                    "end_time": block.end_time,
                    "start_slot": block.start_slot,
                    "periods": block.periods,
                    "subject_code": block.subject_code,
                    "subject_name": block.subject_name,
                    "type": block.type,
                    "confidence": block.confidence,
                    "confidence_score": block.confidence_score,
                    "confidence_reasons": confidence_reasons_to_jsonable(block.confidence_reasons),
                    "block_kind": block.block_kind,
                    "options": [
                        {
                            "subject_code": option.subject_code,
                            "subject_name": option.subject_name,
                            "type": option.type,
                            "place": option.place,
                            "teacher": option.teacher,
                            "confidence": option.confidence,
                            "confidence_score": option.confidence_score,
                            "confidence_reasons": confidence_reasons_to_jsonable(option.confidence_reasons),
                            "raw": option.raw,
                        }
                        for option in block.options
                    ],
                    "raw": block.raw,
                    "bounds": {
                        "start": f"{get_column_letter(block.bounds.min_col)}{block.bounds.min_row}",
                        "end": f"{get_column_letter(block.bounds.max_col)}{block.bounds.max_row}",
                    },
                }
                for block in day_blocks
            ]
            for day, day_blocks in days.items()
        }
        for batch, days in blocks.items()
    }


def confidence_reasons_to_jsonable(reasons: Sequence[ConfidenceReason]) -> list[dict[str, object]]:
    return [
        {
            "code": reason.code.value,
            "penalty": reason.penalty,
            "detail": reason.detail,
        }
        for reason in reasons
    ]
